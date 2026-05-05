import csv
import io
import sqlite3
import os
import re
import secrets
from datetime import datetime, timedelta, date
from functools import wraps
from flask import (
    Flask, render_template, request, redirect, url_for, session, jsonify,
    flash, send_from_directory, abort,
)
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename
from flask.json.provider import DefaultJSONProvider

try:
    from zoneinfo import ZoneInfo  # Python 3.9+
    BR_TZ = ZoneInfo('America/Sao_Paulo')
except Exception:  # pragma: no cover — fallback se zoneinfo indisponível
    BR_TZ = None


def hoje_brasil():
    """Retorna a data de hoje no fuso America/Sao_Paulo (ISO yyyy-mm-dd)."""
    if BR_TZ is not None:
        return datetime.now(BR_TZ).date().isoformat()
    return date.today().isoformat()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'planetun-monitoria-secret-key-2024')
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024  # 25 MB por requisição

DB_PATH = os.path.join(os.path.dirname(__file__), 'monitoria.db')
UPLOAD_ROOT = os.path.join(os.path.dirname(__file__), 'uploads', 'monitorias')
ALLOWED_ANEXO_EXTS = {
    '.pdf', '.png', '.jpg', '.jpeg', '.gif', '.webp', '.heic',
    '.docx', '.xlsx', '.csv', '.txt', '.mp4', '.mov',
}

# Monitoring items structure
ITEMS = {
    1: {'numero': 1, 'categoria': 'GRAVÍSSIMA', 'nome': 'Apontamentos Inadequados', 'descricao': 'Deixou de fazer os devidos apontamentos da análise do risco no laudo...'},
    2: {'numero': 2, 'categoria': 'GRAVÍSSIMA', 'nome': 'Conferência de Informações e fotos do Veículo', 'descricao': 'Não conferiu os dados (modelo, placa e chassi)...'},
    3: {'numero': 3, 'categoria': 'GRAVÍSSIMA', 'nome': 'Validação de Análise por IA', 'descricao': 'Não realizou a conferência dos apontamentos sugeridos pela IA...'},
    4: {'numero': 4, 'categoria': 'GRAVE', 'nome': 'Solicitação Indevida de Fotos', 'descricao': 'Solicitou correções ou envio de fotos adicionais de maneira incorreta...'},
    5: {'numero': 5, 'categoria': 'GRAVE', 'nome': 'Orçamentação Incorreta (aplicável a EA)', 'descricao': 'Realizou orçamento incorreto, desconsiderando imagens, peças...'},
    6: {'numero': 6, 'categoria': 'LEVE', 'nome': 'Preenchimento Incorreto de Dados', 'descricao': 'Erro no preenchimento das informações da vistoria...'},
    7: {'numero': 7, 'categoria': 'LEVE', 'nome': 'Falta de Registro de Solicitação de Fotos e/ou observações', 'descricao': 'Não houve registro sobre solicitação de fotos adicionais...'},
    8: {'numero': 8, 'categoria': 'LEVE', 'nome': 'Relato de Avarias', 'descricao': 'Não realizou o relato das avarias conforme as regras...'},
}

def get_db():
    """Get database connection."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_schema():
    """
    Idempotent schema migration. Safe to call on every startup.
    Adds the 'produtos' table and the 'produto_id' column on 'monitorias'
    when they're missing — keeps Render and local DBs in sync without
    requiring manual migration steps.
    """
    if not os.path.exists(DB_PATH) or os.path.getsize(DB_PATH) == 0:
        # Banco ainda não inicializado — init_db.py é responsável pela criação.
        return
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='monitorias'"
    )
    if cursor.fetchone() is None:
        # Schema base ainda não criado — pula migração.
        conn.close()
        return
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS produtos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE,
            ativo BOOLEAN NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute("PRAGMA table_info(monitorias)")
    cols = {row['name'] for row in cursor.fetchall()}
    if 'produto_id' not in cols:
        cursor.execute(
            'ALTER TABLE monitorias ADD COLUMN produto_id INTEGER REFERENCES produtos(id)'
        )
    # Observação por item
    cursor.execute("PRAGMA table_info(monitoria_itens)")
    item_cols = {row['name'] for row in cursor.fetchall()}
    if 'observacao' not in item_cols:
        cursor.execute('ALTER TABLE monitoria_itens ADD COLUMN observacao TEXT')
    # Tabela de anexos
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS monitoria_anexos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            monitoria_id INTEGER NOT NULL,
            item_numero INTEGER NOT NULL,
            nome_original TEXT NOT NULL,
            nome_arquivo TEXT NOT NULL,
            tamanho INTEGER,
            mime_type TEXT,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (monitoria_id) REFERENCES monitorias(id) ON DELETE CASCADE
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS monitoria_replicas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            monitoria_id INTEGER NOT NULL,
            autor_id INTEGER NOT NULL,
            autor_tipo TEXT NOT NULL CHECK(autor_tipo IN ('perito', 'supervisor')),
            decisao TEXT NOT NULL CHECK(decisao IN ('concordo', 'nao_concordo')),
            justificativa TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (monitoria_id) REFERENCES monitorias(id) ON DELETE CASCADE,
            FOREIGN KEY (autor_id) REFERENCES usuarios(id),
            UNIQUE(monitoria_id, autor_tipo)
        )
    ''')
    conn.commit()
    conn.close()


# Roda a migração ao importar o módulo (cobre Flask CLI, gunicorn e dev server).
ensure_schema()

def login_required(f):
    """Decorator to require login."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def get_usuario(usuario_id):
    """Get user by ID."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM usuarios WHERE id = ?', (usuario_id,))
    usuario = cursor.fetchone()
    conn.close()
    return usuario

def calculate_score(marcados):
    """
    Calculate monitoring score based on marked items.

    Logic:
    - Start at 100 points
    - Items 1-3 are GRAVÍSSIMA: if ANY marked, score = 0
    - Items 4-5 are GRAVE: each marked deducts 60 points (min 0)
    - Items 6-8 are LEVE: each marked deducts 40 points (min 0)
    """
    score = 100

    # Check gravíssima (items 1-3)
    if any(marcados.get(i, False) for i in [1, 2, 3]):
        return 0

    # Deduct for grave failures (items 4-5)
    grave_count = sum(1 for i in [4, 5] if marcados.get(i, False))
    score -= grave_count * 60

    # Deduct for leve failures (items 6-8)
    leve_count = sum(1 for i in [6, 7, 8] if marcados.get(i, False))
    score -= leve_count * 40

    return max(0, score)

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page."""
    if request.method == 'POST':
        email = request.form.get('email')
        senha = request.form.get('senha')

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM usuarios WHERE email = ? AND ativo = 1', (email,))
        usuario = cursor.fetchone()
        conn.close()

        if usuario and check_password_hash(usuario['senha'], senha):
            session['usuario_id'] = usuario['id']
            session['usuario_nome'] = usuario['nome']
            session['usuario_perfil'] = usuario['perfil']
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', erro='Email ou senha inválidos')

    return render_template('login.html')

@app.route('/logout')
def logout():
    """Logout."""
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def index():
    """Redirect to dashboard."""
    if 'usuario_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/formulario', methods=['GET', 'POST'])
@login_required
def formulario():
    """Show monitoring form (supervisors only)."""
    usuario = get_usuario(session['usuario_id'])
    if usuario['perfil'] != 'supervisor':
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        # Get form data
        colaborador_id = request.form.get('colaborador_id', type=int)
        cliente_id = request.form.get('cliente_id', type=int)
        produto_id = request.form.get('produto_id', type=int)
        data_monitoria = request.form.get('data_monitoria')
        data_tratativa = request.form.get('data_tratativa') or None
        data_feedback = request.form.get('data_feedback') or None
        numero_processo = request.form.get('numero_processo') or None
        observacoes = request.form.get('observacoes') or None

        # Get marked items + per-item observation
        marcados = {}
        observacoes_item = {}
        for i in range(1, 9):
            marcados[i] = request.form.get(f'item_{i}') == 'on'
            obs = (request.form.get(f'obs_{i}') or '').strip()
            observacoes_item[i] = obs if (marcados[i] and obs) else None

        # Calculate score
        nota_final = calculate_score(marcados)

        # Save to database
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO monitorias
            (data_monitoria, data_tratativa, data_feedback, colaborador_id, avaliador_id,
             cliente_id, produto_id, numero_processo, observacoes, nota_final)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (data_monitoria, data_tratativa, data_feedback, colaborador_id,
              session['usuario_id'], cliente_id, produto_id, numero_processo,
              observacoes, nota_final))

        monitoria_id = cursor.lastrowid

        # Insert items (com observação)
        for item_numero, marcado in marcados.items():
            cursor.execute('''
                INSERT INTO monitoria_itens (monitoria_id, item_numero, marcado, observacao)
                VALUES (?, ?, ?, ?)
            ''', (monitoria_id, item_numero, marcado, observacoes_item[item_numero]))

        # Anexos por item — só processa quando o item foi marcado
        anexos_dir = os.path.join(UPLOAD_ROOT, str(monitoria_id))
        for item_numero, marcado in marcados.items():
            if not marcado:
                continue
            uploaded = request.files.getlist(f'arquivos_{item_numero}')
            for fobj in uploaded:
                if not fobj or not fobj.filename:
                    continue
                ext = os.path.splitext(fobj.filename)[1].lower()
                if ext not in ALLOWED_ANEXO_EXTS:
                    flash(f'Arquivo "{fobj.filename}" ignorado: extensão não permitida.', 'error')
                    continue
                os.makedirs(anexos_dir, exist_ok=True)
                safe_base = secure_filename(os.path.splitext(fobj.filename)[0]) or 'anexo'
                final_name = f'item{item_numero}_{secrets.token_hex(4)}_{safe_base}{ext}'
                final_path = os.path.join(anexos_dir, final_name)
                fobj.save(final_path)
                tamanho = os.path.getsize(final_path)
                cursor.execute('''
                    INSERT INTO monitoria_anexos
                    (monitoria_id, item_numero, nome_original, nome_arquivo, tamanho, mime_type)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (monitoria_id, item_numero, fobj.filename, final_name,
                      tamanho, fobj.mimetype))

        conn.commit()
        conn.close()

        flash('Monitoria registrada com sucesso.', 'success')
        return redirect(url_for('dashboard'))

    # Get data for dropdowns
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT id, nome FROM usuarios WHERE perfil = "perito" AND ativo = 1 ORDER BY nome')
    peritos = cursor.fetchall()

    cursor.execute('SELECT id, nome FROM clientes ORDER BY nome')
    clientes = cursor.fetchall()

    cursor.execute('SELECT id, nome FROM produtos WHERE ativo = 1 ORDER BY nome')
    produtos_ativos = cursor.fetchall()

    conn.close()

    return render_template('formulario.html',
                         items=ITEMS,
                         peritos=peritos,
                         clientes=clientes,
                         produtos=produtos_ativos,
                         avaliador_nome=usuario['nome'],
                         data_hoje=hoje_brasil())

# ---------------------------------------------------------------------------
# Produtos — gestão administrativa (apenas supervisor)
# ---------------------------------------------------------------------------
ALLOWED_PRODUTO_EXTS = {'.csv', '.xlsx'}


def _parse_produtos_csv(stream):
    """Lê produtos de um stream CSV. Aceita 1ª coluna = nome; ignora header se existir."""
    text = stream.read()
    if isinstance(text, bytes):
        for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                text = text.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = text.decode('utf-8', errors='ignore')
    reader = csv.reader(io.StringIO(text))
    nomes = []
    for i, row in enumerate(reader):
        if not row:
            continue
        nome = (row[0] or '').strip()
        if not nome:
            continue
        # Pula header óbvio na primeira linha
        if i == 0 and nome.lower() in {'produto', 'produtos', 'nome'}:
            continue
        nomes.append(nome)
    return nomes


def _parse_produtos_xlsx(stream):
    """Lê produtos de um stream XLSX (1ª coluna da 1ª aba)."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=stream, read_only=True, data_only=True)
    ws = wb.active
    nomes = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            continue
        cell = row[0]
        nome = (str(cell).strip() if cell is not None else '')
        if not nome:
            continue
        if i == 0 and nome.lower() in {'produto', 'produtos', 'nome'}:
            continue
        nomes.append(nome)
    return nomes


@app.route('/produtos')
@login_required
def produtos():
    """Listagem e gestão de produtos (somente supervisor)."""
    usuario = get_usuario(session['usuario_id'])
    if usuario['perfil'] != 'supervisor':
        return redirect(url_for('dashboard'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT p.id, p.nome, p.ativo, p.created_at,
               (SELECT COUNT(*) FROM monitorias m WHERE m.produto_id = p.id) AS uso
        FROM produtos p
        ORDER BY p.ativo DESC, p.nome ASC
    ''')
    lista = cursor.fetchall()
    conn.close()
    return render_template('produtos.html', produtos=lista)


@app.route('/produtos/novo', methods=['POST'])
@login_required
def produtos_novo():
    """Adiciona um produto manualmente."""
    usuario = get_usuario(session['usuario_id'])
    if usuario['perfil'] != 'supervisor':
        return redirect(url_for('dashboard'))
    nome = (request.form.get('nome') or '').strip()
    if not nome:
        flash('Informe o nome do produto.', 'error')
        return redirect(url_for('produtos'))
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO produtos (nome) VALUES (?)', (nome,))
        conn.commit()
        flash(f'Produto "{nome}" cadastrado.', 'success')
    except sqlite3.IntegrityError:
        flash(f'Produto "{nome}" já existe.', 'error')
    finally:
        conn.close()
    return redirect(url_for('produtos'))


@app.route('/produtos/upload', methods=['POST'])
@login_required
def produtos_upload():
    """Upload em lote (CSV ou XLSX). 1ª coluna = nome do produto."""
    usuario = get_usuario(session['usuario_id'])
    if usuario['perfil'] != 'supervisor':
        return redirect(url_for('dashboard'))
    file = request.files.get('arquivo')
    if not file or not file.filename:
        flash('Selecione um arquivo CSV ou XLSX.', 'error')
        return redirect(url_for('produtos'))
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_PRODUTO_EXTS:
        flash('Formato não suportado. Use .csv ou .xlsx.', 'error')
        return redirect(url_for('produtos'))
    try:
        if ext == '.csv':
            nomes = _parse_produtos_csv(file.stream)
        else:
            nomes = _parse_produtos_xlsx(file.stream)
    except Exception as exc:
        flash(f'Erro ao ler arquivo: {exc}', 'error')
        return redirect(url_for('produtos'))
    if not nomes:
        flash('Nenhum nome de produto encontrado no arquivo.', 'error')
        return redirect(url_for('produtos'))
    conn = get_db()
    cursor = conn.cursor()
    inseridos, duplicados = 0, 0
    for nome in nomes:
        try:
            cursor.execute('INSERT INTO produtos (nome) VALUES (?)', (nome,))
            inseridos += 1
        except sqlite3.IntegrityError:
            duplicados += 1
    conn.commit()
    conn.close()
    msg = f'{inseridos} produto(s) cadastrado(s).'
    if duplicados:
        msg += f' {duplicados} já existia(m) e foram ignorados.'
    flash(msg, 'success')
    return redirect(url_for('produtos'))


@app.route('/produtos/<int:produto_id>/toggle', methods=['POST'])
@login_required
def produtos_toggle(produto_id):
    """Ativa ou desativa um produto (sem deletar — preserva histórico)."""
    usuario = get_usuario(session['usuario_id'])
    if usuario['perfil'] != 'supervisor':
        return redirect(url_for('dashboard'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE produtos SET ativo = 1 - ativo WHERE id = ?', (produto_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('produtos'))


# ---------------------------------------------------------------------------
# Histórico — lista completa de monitorias com filtros
# ---------------------------------------------------------------------------
@app.route('/historico')
@login_required
def historico():
    usuario = get_usuario(session['usuario_id'])
    conn = get_db()
    cursor = conn.cursor()

    perito_id = request.args.get('perito_id', type=int)
    cliente_id = request.args.get('cliente_id', type=int)
    produto_id = request.args.get('produto_id', type=int)
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)

    query = '''
        SELECT m.id, m.data_monitoria, m.nota_final, m.numero_processo,
               u1.nome AS perito_nome, u2.nome AS avaliador_nome,
               c.nome AS cliente_nome,
               COALESCE(p.nome, '—') AS produto_nome
        FROM monitorias m
        JOIN usuarios u1 ON m.colaborador_id = u1.id
        JOIN usuarios u2 ON m.avaliador_id = u2.id
        JOIN clientes c ON m.cliente_id = c.id
        LEFT JOIN produtos p ON m.produto_id = p.id
        WHERE 1=1
    '''
    params = []
    if usuario['perfil'] == 'perito':
        query += ' AND m.colaborador_id = ?'
        params.append(usuario['id'])
    elif perito_id:
        query += ' AND m.colaborador_id = ?'
        params.append(perito_id)
    if cliente_id:
        query += ' AND m.cliente_id = ?'
        params.append(cliente_id)
    if produto_id:
        query += ' AND m.produto_id = ?'
        params.append(produto_id)
    if mes and ano:
        query += ' AND strftime("%m", m.data_monitoria) = ? AND strftime("%Y", m.data_monitoria) = ?'
        params.extend([f'{mes:02d}', f'{ano:04d}'])
    query += ' ORDER BY m.data_monitoria DESC, m.id DESC'

    cursor.execute(query, params)
    monitorias = cursor.fetchall()

    cursor.execute('SELECT id, nome FROM usuarios WHERE perfil = "perito" AND ativo = 1 ORDER BY nome')
    peritos_list = cursor.fetchall()
    cursor.execute('SELECT id, nome FROM clientes ORDER BY nome')
    clientes_list = cursor.fetchall()
    cursor.execute('SELECT id, nome FROM produtos WHERE ativo = 1 ORDER BY nome')
    produtos_list = cursor.fetchall()
    conn.close()

    return render_template(
        'historico.html',
        monitorias=monitorias,
        peritos=peritos_list,
        clientes=clientes_list,
        produtos=produtos_list,
        filtros={
            'perito_id': perito_id, 'cliente_id': cliente_id,
            'produto_id': produto_id, 'mes': mes, 'ano': ano,
        },
        is_supervisor=(usuario['perfil'] == 'supervisor'),
    )


# ---------------------------------------------------------------------------
# Peritos — ranking e desempenho individual (apenas supervisor)
# ---------------------------------------------------------------------------
@app.route('/peritos')
@login_required
def peritos():
    usuario = get_usuario(session['usuario_id'])
    if usuario['perfil'] != 'supervisor':
        return redirect(url_for('dashboard'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT u.id, u.nome, u.email, u.ativo,
               COUNT(m.id) AS total_monitorias,
               COALESCE(AVG(m.nota_final), 0) AS nota_media,
               MAX(m.data_monitoria) AS ultima_data
        FROM usuarios u
        LEFT JOIN monitorias m ON m.colaborador_id = u.id
        WHERE u.perfil = 'perito'
        GROUP BY u.id, u.nome, u.email, u.ativo
        ORDER BY nota_media DESC, u.nome ASC
    ''')
    lista = cursor.fetchall()
    conn.close()
    return render_template('peritos.html', peritos=lista)


# ---------------------------------------------------------------------------
# Servir anexos de monitoria (login obrigatório, perito só vê os seus)
# ---------------------------------------------------------------------------
@app.route('/uploads/monitorias/<int:monitoria_id>/<path:filename>')
@login_required
def servir_anexo(monitoria_id, filename):
    usuario = get_usuario(session['usuario_id'])
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT colaborador_id FROM monitorias WHERE id = ?', (monitoria_id,))
    row = cursor.fetchone()
    conn.close()
    if row is None:
        abort(404)
    if usuario['perfil'] == 'perito' and row['colaborador_id'] != usuario['id']:
        abort(403)
    safe_dir = os.path.join(UPLOAD_ROOT, str(monitoria_id))
    return send_from_directory(safe_dir, filename, as_attachment=False)


@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard page."""
    usuario = get_usuario(session['usuario_id'])

    conn = get_db()
    cursor = conn.cursor()

    # Get peritos for filter
    cursor.execute('SELECT id, nome FROM usuarios WHERE perfil = "perito" AND ativo = 1 ORDER BY nome')
    peritos = cursor.fetchall()

    # Get supervisors for filter
    cursor.execute('SELECT id, nome FROM usuarios WHERE perfil = "supervisor" AND ativo = 1 ORDER BY nome')
    supervisores = cursor.fetchall()

    # Get produtos ativos for filter
    cursor.execute('SELECT id, nome FROM produtos WHERE ativo = 1 ORDER BY nome')
    produtos_ativos = cursor.fetchall()

    conn.close()

    return render_template('dashboard.html',
                         usuario=usuario,
                         peritos=peritos,
                         supervisores=supervisores,
                         produtos=produtos_ativos)

@app.route('/monitoria/<int:monitoria_id>')
@login_required
def monitoria_detalhe(monitoria_id):
    """Show details of a single monitoria."""
    usuario = get_usuario(session['usuario_id'])

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT m.*, u1.nome as colaborador_nome, u2.nome as avaliador_nome, c.nome as cliente_nome
        FROM monitorias m
        JOIN usuarios u1 ON m.colaborador_id = u1.id
        JOIN usuarios u2 ON m.avaliador_id = u2.id
        JOIN clientes c ON m.cliente_id = c.id
        WHERE m.id = ?
    ''', (monitoria_id,))
    monitoria = cursor.fetchone()

    if monitoria is None:
        conn.close()
        return render_template('monitoria_detalhe.html', monitoria=None, itens_marcados=[], items=ITEMS), 404

    # Perito só pode ver as próprias monitorias
    if usuario['perfil'] == 'perito' and monitoria['colaborador_id'] != usuario['id']:
        conn.close()
        return redirect(url_for('dashboard'))

    cursor.execute('''
        SELECT item_numero, marcado, observacao FROM monitoria_itens
        WHERE monitoria_id = ?
        ORDER BY item_numero
    ''', (monitoria_id,))
    itens_rows = cursor.fetchall()

    cursor.execute('''
        SELECT id, item_numero, nome_original, nome_arquivo, tamanho, mime_type, uploaded_at
        FROM monitoria_anexos
        WHERE monitoria_id = ?
        ORDER BY item_numero, uploaded_at
    ''', (monitoria_id,))
    anexos_rows = cursor.fetchall()

    cursor.execute('''
        SELECT r.*, u.nome AS autor_nome
        FROM monitoria_replicas r
        JOIN usuarios u ON r.autor_id = u.id
        WHERE r.monitoria_id = ?
        ORDER BY r.criado_em
    ''', (monitoria_id,))
    replicas = cursor.fetchall()
    conn.close()

    # Mapas auxiliares
    marcados_map = {row['item_numero']: bool(row['marcado']) for row in itens_rows}
    obs_map = {row['item_numero']: (row['observacao'] or '') for row in itens_rows}
    anexos_por_item = {}
    for a in anexos_rows:
        anexos_por_item.setdefault(a['item_numero'], []).append(a)

    replica_perito = next((r for r in replicas if r['autor_tipo'] == 'perito'), None)
    replica_supervisor = next((r for r in replicas if r['autor_tipo'] == 'supervisor'), None)

    # Permissões da UI de réplica
    eh_perito_avaliado = (
        usuario['perfil'] == 'perito'
        and monitoria['colaborador_id'] == usuario['id']
    )
    eh_supervisor = (usuario['perfil'] == 'supervisor')

    pode_replicar_perito = eh_perito_avaliado and replica_perito is None
    pode_replicar_supervisor = (
        eh_supervisor and replica_perito is not None and replica_supervisor is None
    )

    return render_template('monitoria_detalhe.html',
                           monitoria=monitoria,
                           marcados_map=marcados_map,
                           obs_map=obs_map,
                           anexos_por_item=anexos_por_item,
                           items=ITEMS,
                           usuario=usuario,
                           replica_perito=replica_perito,
                           replica_supervisor=replica_supervisor,
                           pode_replicar_perito=pode_replicar_perito,
                           pode_replicar_supervisor=pode_replicar_supervisor,
                           eh_perito_avaliado=eh_perito_avaliado)

# ---------------------------------------------------------------------------
# Réplica perito ↔ supervisor (UI funcional, sem e-mail/tokens — Pendência 2)
# ---------------------------------------------------------------------------
@app.route('/monitoria/<int:monitoria_id>/replica', methods=['POST'])
@login_required
def monitoria_replica(monitoria_id):
    usuario = get_usuario(session['usuario_id'])
    decisao = (request.form.get('decisao') or '').strip()
    justificativa = (request.form.get('justificativa') or '').strip() or None

    if decisao not in {'concordo', 'nao_concordo'}:
        flash('Decisão inválida.', 'error')
        return redirect(url_for('monitoria_detalhe', monitoria_id=monitoria_id))

    if decisao == 'nao_concordo' and not justificativa:
        flash('Justificativa é obrigatória quando você marca "Não concordo".', 'error')
        return redirect(url_for('monitoria_detalhe', monitoria_id=monitoria_id))

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT colaborador_id FROM monitorias WHERE id = ?', (monitoria_id,))
    monitoria = cursor.fetchone()
    if monitoria is None:
        conn.close()
        abort(404)

    # Quem tem permissão para esta réplica?
    if usuario['perfil'] == 'perito':
        if monitoria['colaborador_id'] != usuario['id']:
            conn.close()
            abort(403)
        autor_tipo = 'perito'
    elif usuario['perfil'] == 'supervisor':
        # Supervisor só pode responder DEPOIS que o perito tiver replicado
        cursor.execute(
            'SELECT 1 FROM monitoria_replicas WHERE monitoria_id = ? AND autor_tipo = "perito"',
            (monitoria_id,))
        if cursor.fetchone() is None:
            conn.close()
            flash('O supervisor só pode responder depois que o perito enviar a manifestação.', 'error')
            return redirect(url_for('monitoria_detalhe', monitoria_id=monitoria_id))
        autor_tipo = 'supervisor'
    else:
        conn.close()
        abort(403)

    try:
        cursor.execute('''
            INSERT INTO monitoria_replicas
            (monitoria_id, autor_id, autor_tipo, decisao, justificativa)
            VALUES (?, ?, ?, ?, ?)
        ''', (monitoria_id, usuario['id'], autor_tipo, decisao, justificativa))
        conn.commit()
        flash('Manifestação registrada.', 'success')
    except sqlite3.IntegrityError:
        flash('Você já registrou sua manifestação para esta monitoria.', 'error')
    finally:
        conn.close()

    return redirect(url_for('monitoria_detalhe', monitoria_id=monitoria_id))


@app.route('/api/dashboard-data')
@login_required
def api_dashboard_data():
    """API endpoint for dashboard data with filters."""
    usuario = get_usuario(session['usuario_id'])

    # Get filter parameters
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)
    perito_id = request.args.get('perito_id', type=int)
    supervisor_id = request.args.get('supervisor_id', type=int)
    produto_id = request.args.get('produto_id', type=int)

    conn = get_db()
    cursor = conn.cursor()

    # Build base query
    query = '''
        SELECT m.*, u1.nome as colaborador_nome, u2.nome as avaliador_nome, c.nome as cliente_nome
        FROM monitorias m
        JOIN usuarios u1 ON m.colaborador_id = u1.id
        JOIN usuarios u2 ON m.avaliador_id = u2.id
        JOIN clientes c ON m.cliente_id = c.id
        WHERE 1=1
    '''
    params = []

    # If perito, only show their own data
    if usuario['perfil'] == 'perito':
        query += ' AND m.colaborador_id = ?'
        params.append(usuario['id'])
    elif perito_id:
        query += ' AND m.colaborador_id = ?'
        params.append(perito_id)

    if supervisor_id:
        query += ' AND m.avaliador_id = ?'
        params.append(supervisor_id)

    if produto_id:
        query += ' AND m.produto_id = ?'
        params.append(produto_id)

    if mes and ano:
        query += ' AND strftime("%m", m.data_monitoria) = ? AND strftime("%Y", m.data_monitoria) = ?'
        params.extend([f'{mes:02d}', f'{ano:04d}'])

    query += ' ORDER BY m.data_monitoria DESC'

    cursor.execute(query, params)
    monitorias = cursor.fetchall()

    # Calculate KPIs
    total_monitorias = len(monitorias)

    if total_monitorias > 0:
        nota_media = sum(m['nota_final'] for m in monitorias) / total_monitorias
        conformidade = sum(1 for m in monitorias if m['nota_final'] == 100) / total_monitorias * 100

        # Count gravíssima failures
        gravissima_count = 0
        for m in monitorias:
            cursor.execute('''
                SELECT COUNT(*) as count FROM monitoria_itens
                WHERE monitoria_id = ? AND item_numero IN (1, 2, 3) AND marcado = 1
            ''', (m['id'],))
            gravissima_count += cursor.fetchone()['count']

        pct_gravissima = (gravissima_count / total_monitorias) * 100 if total_monitorias > 0 else 0
    else:
        nota_media = 0
        conformidade = 0
        pct_gravissima = 0

    # Get data by perito
    cursor.execute('''
        SELECT u.nome, AVG(m.nota_final) as media_nota, COUNT(m.id) as total
        FROM monitorias m
        JOIN usuarios u ON m.colaborador_id = u.id
        WHERE u.perfil = 'perito'
    ''')

    if usuario['perfil'] == 'perito':
        cursor.execute('''
            SELECT u.nome, AVG(m.nota_final) as media_nota, COUNT(m.id) as total
            FROM monitorias m
            JOIN usuarios u ON m.colaborador_id = u.id
            WHERE u.perfil = 'perito' AND m.colaborador_id = ?
        ''', (usuario['id'],))
    else:
        cursor.execute('''
            SELECT u.nome, AVG(m.nota_final) as media_nota, COUNT(m.id) as total
            FROM monitorias m
            JOIN usuarios u ON m.colaborador_id = u.id
            WHERE u.perfil = 'perito'
        ''')
        if perito_id:
            cursor.execute('''
                SELECT u.nome, AVG(m.nota_final) as media_nota, COUNT(m.id) as total
                FROM monitorias m
                JOIN usuarios u ON m.colaborador_id = u.id
                WHERE u.perfil = 'perito' AND m.colaborador_id = ?
            ''', (perito_id,))

    peritos_stats = cursor.fetchall()
    peritos_data = {
        'labels': [p['nome'] for p in peritos_stats],
        'data': [round(p['media_nota'], 2) if p['media_nota'] else 0 for p in peritos_stats]
    }

    # Get monthly evolution data (last 12 months)
    cursor.execute('''
        SELECT strftime("%Y-%m", m.data_monitoria) as mes_ano, AVG(m.nota_final) as media
        FROM monitorias m
        WHERE m.data_monitoria >= date('now', '-12 months')
        GROUP BY strftime("%Y-%m", m.data_monitoria)
        ORDER BY mes_ano
    ''')

    monthly_stats = cursor.fetchall()
    monthly_data = {
        'labels': [m['mes_ano'] for m in monthly_stats],
        'data': [round(m['media'], 2) if m['media'] else 0 for m in monthly_stats]
    }

    # Failure distribution
    sem_falha = sum(1 for m in monitorias if m['nota_final'] == 100)
    leve = sum(1 for m in monitorias if m['nota_final'] in range(20, 100))
    grave = sum(1 for m in monitorias if m['nota_final'] in range(1, 20))
    gravissima = sum(1 for m in monitorias if m['nota_final'] == 0)

    failure_data = {
        'labels': ['Sem Falha', 'Falha Leve', 'Falha Grave', 'Falha Gravíssima'],
        'data': [sem_falha, leve, grave, gravissima]
    }

    # Last monitorias
    ultimas_monitorias = [
        {
            'id': m['id'],
            'data': m['data_monitoria'],
            'perito': m['colaborador_nome'],
            'supervisor': m['avaliador_nome'],
            'cliente': m['cliente_nome'],
            'nota': round(m['nota_final'], 2)
        }
        for m in monitorias[:10]
    ]

    conn.close()

    return jsonify({
        'kpis': {
            'total_monitorias': total_monitorias,
            'nota_media': round(nota_media, 2),
            'pct_gravissima': round(pct_gravissima, 2),
            'pct_conformidade': round(conformidade, 2)
        },
        'peritos': peritos_data,
        'monthly': monthly_data,
        'failures': failure_data,
        'ultimas_monitorias': ultimas_monitorias
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, host='0.0.0.0', port=port)
